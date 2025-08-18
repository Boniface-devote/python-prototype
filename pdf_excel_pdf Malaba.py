from flask import Flask, request, render_template_string, send_file
import pdfplumber
import re
from io import BytesIO
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table

app = Flask(__name__)

# Global variables to store the modified files in memory (for simplicity, assuming single user in debug mode)
modified_excel_global = None
modified_pdf_global = None

# HTML template for the upload form
UPLOAD_FORM = '''
<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>PDF Data Extractor</title>
</head>
<body>
    <h1>Upload PDF MALABA draft feri and Excel Template</h1>
    <form method="post" enctype="multipart/form-data">
        <label>PDF File:</label> <input type="file" name="pdf_file"><br><br>
        <label>Excel File (optional):</label> <input type="file" name="excel_file"><br><br>
        <label>Freight Number:</label> <input type="number" name="freight_number" min="1" placeholder="Enter Freight Number (e.g., 200, 250, 500)"><br><br>
        <label>Container Type:</label><br>
        <input type="checkbox" name="container_type" value="40FT"> 1*40FT<br>
        <input type="checkbox" name="container_type" value="20FT"> 1*20FT<br><br>
        <label>Number of Containers:</label> <input type="number" name="num_containers" min="1" placeholder="Enter Number of Containers (default: 1)"><br><br>
        <input type="submit" value="Upload and Process">
    </form>
    {% if data %}
    <h2>Extracted Data</h2>
    <pre>{{ data | tojson(indent=4) }}</pre>
    {% endif %}
    {% if modified %}
    <h2>Downloads</h2>
    <a href="/download_excel">Download Modified Excel</a><br>
    <a href="/download_pdf">Download as PDF</a>
    {% endif %}
</body>
</html>
'''

def extract_data_from_pdf(pdf_content):
    """
    Extract structured data from the PDF content.
    Supports both FERI and A.D style freight certificates.
    """
    extracted = {}
    with pdfplumber.open(BytesIO(pdf_content)) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text() or ""

    # Clean up text
    text = text.replace('\n', ' ').strip()

    # Use regex to extract key fields based on patterns from the sample PDFs
    # FERI Number or Attestation Number
    match = re.search(r'(?:FERI N°|VALIDATION|A\.D\s+N°)\s*:\s*(\w+)', text)
    if match:
        extracted['feri_number'] = match.group(1)

    # Importateur (only name)
    match = re.search(r'IMPORTATEUR\s*:\s*([^\s;][^;]*?)(?:\s*;|EXPORTATEUR|ADD:|$)', text, re.DOTALL)
    if match:
        importateur_name = match.group(1).strip()
        name_match = re.match(r'^[A-Z\s()]+(?:\s*\([A-Z]+\)\s*[A-Z]+)?', importateur_name)
        importateur_name = name_match.group(0).strip() if name_match else importateur_name
        extracted['importateur'] = importateur_name

    # Transitaire (only name)
    match = re.search(r'TRANSITAIRE\s*:\s*([^\s;][^;]*?)(?:\s*Forwarding agent|DEST\.|ADD:|$)', text, re.DOTALL)
    if match:
        forwarding_name = match.group(1).strip()
        name_match = re.match(r'^[A-Z\s()]+(?:\s*\([A-Z]+\)\s*[A-Z]+)?', forwarding_name)
        extracted['transitaire'] = name_match.group(0).strip() if name_match else forwarding_name

    # BL (Bill of Lading)
    match = re.search(r'(?:BL|TITRE DE TRANSPORT)\s*:\s*(.+?)(?:\s*ARMATEUR|TRANS|$)', text)
    if match:
        extracted['bl'] = match.group(1).strip()

    # CBM (Cubic Meters) - Prioritize the CBM field explicitly near weight or container details
    match = re.search(r'(\d+\.\d+\s*CBM)\s*(?=(?:POIDS|TEU|CONTENEUR|$))', text)
    if match:
        extracted['cbm'] = match.group(1).strip()

    # Gross Weight (optional for A.D style)
    match = re.search(r'POIDS BRUT\s*:\s*([\d\.]+)\s*(?:Kg|T)', text)
    if match:
        extracted['gross_weight'] = float(match.group(1))

    # Exporter (optional for A.D style)
    match = re.search(r'EXPORTATEUR\s*([^\s;][^;]*?)(?:\s*;|TRANSITAIRE|$)', text, re.DOTALL)
    if match:
        exporter_name = match.group(1).strip()
        name_match = re.match(r'^[A-Z\s()]+(?:\s*\([A-Z]+\)\s*[A-Z]+)?', exporter_name)
        exporter_name = name_match.group(0).strip() if name_match else exporter_name
        if exporter_name.endswith(' E'):
            exporter_name = exporter_name[:-2].strip()
        extracted['exporter'] = exporter_name

    return extracted

@app.route('/', methods=['GET', 'POST'])
def upload_pdf():
    data = None
    modified = False
    if request.method == 'POST':
        if 'pdf_file' not in request.files:
            return "No PDF file part"
        pdf_file = request.files['pdf_file']
        if pdf_file.filename == '':
            return "No selected PDF file"
        if pdf_file and pdf_file.filename.endswith('.pdf'):
            pdf_content = pdf_file.read()
            data = extract_data_from_pdf(pdf_content)

            # Check for Excel file
            excel_content = None
            if 'excel_file' in request.files:
                excel_file = request.files['excel_file']
                if excel_file and excel_file.filename.endswith('.xlsx'):
                    excel_content = BytesIO(excel_file.read())

            # Get Freight Number from form input
            freight_number = request.form.get('freight_number', '').strip()
            try:
                freight_number = int(freight_number) if freight_number else None
            except ValueError:
                freight_number = None

            # Get Container Type from form input
            container_type = request.form.get('container_type', '')

            # Get Number of Containers from form input, default to 1
            num_containers = request.form.get('num_containers', '').strip()
            try:
                num_containers = int(num_containers) if num_containers else 1
            except ValueError:
                num_containers = 1

            if excel_content:
                # Load workbook
                wb = openpyxl.load_workbook(excel_content)
                ws = wb.active

                # Insert data into specific cells with required formats
                if 'feri_number' in data:
                    ws['E6'] = f"FERI/AD: {data['feri_number']}"
                    ws['B11'] = f"CERTIFICATE (FERI/ADR/AD) No : {data['feri_number']}"
                if 'transitaire' in data:
                    ws['B8'] = f"DEBTOR: {data['transitaire']}"
                if 'importateur' in data:
                    ws['B10'] = f"IMPORTER: {data['importateur']}"
                if 'bl' in data:
                    ws['B14'] = data['bl']
                if 'cbm' in data:
                    cbm_value = float(data['cbm'].replace(' CBM', ''))
                    ws['D14'] = cbm_value
                if container_type == '40FT':
                    ws['D14'] = 110
                elif container_type == '20FT':
                    ws['D14'] = 60
                # Insert number of containers into E14 and D17
                ws['E14'] = num_containers
                ws['D17'] = num_containers
                # Calculate and insert number of containers * 250 into D18
                ws['D18'] = num_containers * 250
                if freight_number is not None:
                    ws['D18'] = freight_number  # Freight number overrides if provided

                # Define the order of fields to append as a new row
                fields = ['feri_number', 'exporter', 'transitaire', 'bl', 'cbm', 'gross_weight']
                row_data = [data.get(f, '') for f in fields]
                ws.append(row_data)

                # Save modified Excel
                modified_excel = BytesIO()
                wb.save(modified_excel)
                modified_excel.seek(0)

                # Create PDF from the Excel data
                sheet_data = [list(row) for row in ws.iter_rows(values_only=True)]
                pdf_buffer = BytesIO()
                doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
                table = Table(sheet_data)
                doc.build([table])
                pdf_buffer.seek(0)

                # Store in globals
                global modified_excel_global, modified_pdf_global
                modified_excel_global = modified_excel
                modified_pdf_global = pdf_buffer

                modified = True

    return render_template_string(UPLOAD_FORM, data=data, modified=modified)

@app.route('/download_excel')
def download_excel():
    global modified_excel_global
    if modified_excel_global is None:
        return "No modified Excel available"
    modified_excel_global.seek(0)
    return send_file(modified_excel_global, as_attachment=True, download_name='modified.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/download_pdf')
def download_pdf():
    global modified_pdf_global
    if modified_pdf_global is None:
        return "No PDF available"
    modified_pdf_global.seek(0)
    return send_file(modified_pdf_global, as_attachment=True, download_name='modified.pdf', mimetype='application/pdf')

if __name__ == '__main__':
    app.run(debug=True)