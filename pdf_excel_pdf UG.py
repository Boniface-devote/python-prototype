from flask import Flask, request, render_template_string, send_file
import pdfplumber
import re
from io import BytesIO
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table

app = Flask(__name__)

# Global variables to store the modified files in memory (for simplicity, assuming single user in debug mode)
modified_excel_global = None
modified_pdf_global = None

# HTML template for the upload form
UPLOAD_FORM = '''
<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>PDF Data Extractor</title>
</head>
<body>
    <h1>Upload PDF Freight Certificate and Optional Excel Template</h1>
    <form method="post" enctype="multipart/form-data">
        <label>PDF File:</label> <input type="file" name="pdf_file"><br><br>
        <label>Excel File (optional):</label> <input type="file" name="excel_file"><br><br>
        <label>Freight Number:</label> <input type="number" name="freight_number" min="1" placeholder="Enter Freight Number (e.g., 200, 250, 500)"><br><br>
        <input type="submit" value="Upload and Process">
    </form>
    {% if data %}
    <h2>Extracted Data</h2>
    <pre>{{ data | tojson(indent=4) }}</pre>
    {% endif %}
    {% if modified %}
    <h2>Downloads</h2>
    <a href="/download_excel">Download Modified Excel</a><br>
    <a href="/download_pdf">Download as PDF</a>
    {% endif %}
</body>
</html>
'''

def extract_data_from_pdf(pdf_content):
    """
    Extract structured data from the PDF content.
    Assumes the PDF structure is similar to the provided freight certificate.
    """
    extracted = {}
    with pdfplumber.open(BytesIO(pdf_content)) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text() or ""

    # Clean up text
    text = text.replace('\n', ' ').strip()

    # Use regex to extract key fields based on patterns from the sample PDF
    # Attestation Number
    match = re.search(r'A\.D\s+N°\s*(\w+)', text)
    if match:
        extracted['attestation_number'] = match.group(1)

    # Importateur (only name)
    match = re.search(r'IMPORTATEUR\s*:\s*([^\s;][^;]*?)(?:\s*;|EXPORTATEUR)', text, re.DOTALL)
    if match:
        importateur_name = match.group(1).strip()
        name_match = re.match(r'^[A-Z\s()]+(?:\s*\([A-Z]+\)\s*[A-Z]+)?', importateur_name)
        importateur_name = name_match.group(0).strip() if name_match else importateur_name
        extracted['importateur'] = importateur_name

    # Exporter (only name)
    match = re.search(r'EXPORTATEUR\s*([^\s;][^;]*?)(?:\s*;|TRANSITAIRE)', text, re.DOTALL)
    if match:
        exporter_name = match.group(1).strip()
        # Extract only the name by taking text before any address-like patterns
        name_match = re.match(r'^[A-Z\s()]+(?:\s*\([A-Z]+\)\s*[A-Z]+)?', exporter_name)
        exporter_name = name_match.group(0).strip() if name_match else exporter_name
        # Remove trailing " E" if present
        if exporter_name.endswith(' E'):
            exporter_name = exporter_name[:-2].strip()
        extracted['exporter'] = exporter_name
        
    # Forwarding Agent (only name)
    match = re.search(r'TRANSITAIRE\s*:\s*([^\s;][^;]*?)(?:\s*Forwarding agent|DEST\.)', text, re.DOTALL)
    if match:
        forwarding_name = match.group(1).strip()
        # Extract only the name by taking text before any additional details
        name_match = re.match(r'^[A-Z\s()]+(?:\s*\([A-Z]+\)\s*[A-Z]+)?', forwarding_name)
        extracted['forwarding_agent'] = name_match.group(0).strip() if name_match else forwarding_name

    # Transport ID
    match = re.search(r'TITRE DE TRANSPORT\s*:\s*(.+?)\s*TRANS', text)
    if match:
        extracted['transport_id'] = match.group(1).strip()

    # CBM (Cubic Meters)
    match = re.search(r'(\d+\.\d+\s*CBM)', text)
    if match:
        extracted['cbm'] = match.group(1).strip()

    # Gross Weight
    match = re.search(r'POIDS BRUT\s*:\s*([\d\.]+)\s*Kg', text)
    if match:
        extracted['gross_weight'] = float(match.group(1))

    return extracted

@app.route('/', methods=['GET', 'POST'])
def upload_pdf():
    data = None
    modified = False
    if request.method == 'POST':
        if 'pdf_file' not in request.files:
            return "No PDF file part"
        pdf_file = request.files['pdf_file']
        if pdf_file.filename == '':
            return "No selected PDF file"
        if pdf_file and pdf_file.filename.endswith('.pdf'):
            pdf_content = pdf_file.read()
            data = extract_data_from_pdf(pdf_content)

            # Check for Excel file
            excel_content = None
            if 'excel_file' in request.files:
                excel_file = request.files['excel_file']
                if excel_file and excel_file.filename.endswith('.xlsx'):
                    excel_content = BytesIO(excel_file.read())

            # Get Freight Number from form input
            freight_number = request.form.get('freight_number', '').strip()
            try:
                freight_number = int(freight_number) if freight_number else None
            except ValueError:
                freight_number = None

            if excel_content:
                # Load workbook
                wb = openpyxl.load_workbook(excel_content)
                ws = wb.active

                # Insert data into specific cells with required formats
                if 'attestation_number' in data:
                    ws['E6'] = f"FERI/AD: {data['attestation_number']}"
                    ws['B11'] = f"CERTIFICATE (FERI/ADR/AD) No : {data['attestation_number']}"
                if 'forwarding_agent' in data:
                    ws['B8'] = f"DEBTOR: {data['forwarding_agent']}"
                if 'importateur' in data:
                    ws['B10'] = f"IMPORTER: {data['importateur']}"
                if 'transport_id' in data:
                    ws['B14'] = data['transport_id']
                if 'cbm' in data:
                    # Extract the numeric part from cbm (e.g., "51.899" from "51.899 CBM")
                    cbm_value = float(data['cbm'].replace(' CBM', ''))
                    ws['D14'] = cbm_value
                if freight_number is not None:
                    ws['D18'] = freight_number

                # Define the order of fields to append as a new row
                fields = ['attestation_number', 'exporter', 'forwarding_agent', 'transport_id', 'cbm', 'gross_weight']
                row_data = [data.get(f, '') for f in fields]
                ws.append(row_data)

                # Save modified Excel
                modified_excel = BytesIO()
                wb.save(modified_excel)
                modified_excel.seek(0)

                # Create PDF from the Excel data
                sheet_data = [list(row) for row in ws.iter_rows(values_only=True)]
                pdf_buffer = BytesIO()
                doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
                table = Table(sheet_data)
                doc.build([table])
                pdf_buffer.seek(0)

                # Store in globals
                global modified_excel_global, modified_pdf_global
                modified_excel_global = modified_excel
                modified_pdf_global = pdf_buffer

                modified = True

    return render_template_string(UPLOAD_FORM, data=data, modified=modified)

@app.route('/download_excel')
def download_excel():
    global modified_excel_global
    if modified_excel_global is None:
        return "No modified Excel available"
    modified_excel_global.seek(0)
    return send_file(modified_excel_global, as_attachment=True, download_name='modified.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/download_pdf')
def download_pdf():
    global modified_pdf_global
    if modified_pdf_global is None:
        return "No PDF available"
    modified_pdf_global.seek(0)
    return send_file(modified_pdf_global, as_attachment=True, download_name='modified.pdf', mimetype='application/pdf')

if __name__ == '__main__':
    app.run(debug=True)