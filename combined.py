from flask import Flask, request, render_template_string, send_file
import pdfplumber
import re
from io import BytesIO
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table

app = Flask(__name__)

# Global variables to store the modified files in memory
modified_excel_global = None
modified_pdf_global = None

# HTML template with tabs for Normal FERI and Maritime PDFs
UPLOAD_FORM = '''
<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>PDF Data Extractor</title>
    <script src="https://cdn.tailwindcss.com"></script>
</head>
<body class="bg-gray-100 font-sans">
    <div class="container mx-auto p-6">
        <h1 class="text-2xl font-bold mb-6 text-center">PDF Data Extractor</h1>
        <!-- Tabs -->
        <div class="mb-4">
            <ul class="flex border-b">
                <li class="mr-1">
                    <button id="normal-tab" class="tab-button bg-white inline-block py-2 px-4 text-blue-600 font-semibold border-b-2 border-blue-600" onclick="showTab('normal')">Normal FERI PDFs</button>
                </li>
                <li class="mr-1">
                    <button id="maritime-tab" class="tab-button bg-white inline-block py-2 px-4 text-gray-600 font-semibold" onclick="showTab('maritime')">Maritime PDFs</button>
                </li>
            </ul>
        </div>
        <!-- Tab Content -->
        <div id="normal-form" class="tab-content">
            <form method="post" enctype="multipart/form-data" action="/process/normal">
                <h2 class="text-xl font-semibold mb-4">Upload Normal FERI PDF and Optional Excel Template</h2>
                <div class="mb-4">
                    <label class="block text-gray-700">PDF File:</label>
                    <input type="file" name="pdf_file" accept=".pdf" class="mt-1 p-2 border rounded w-full" required>
                </div>
                <div class="mb-4">
                    <label class="block text-gray-700">Excel File (optional):</label>
                    <input type="file" name="excel_file" accept=".xlsx" class="mt-1 p-2 border rounded w-full">
                </div>
                <div class="mb-4">
                    <label class="block text-gray-700">Freight Number:</label>
                    <input type="number" name="freight_number" min="1" placeholder="Enter Freight Number (e.g., 200, 250, 500)" class="mt-1 p-2 border rounded w-full">
                </div>
                <input type="submit" value="Upload and Process" class="bg-blue-600 text-white py-2 px-4 rounded hover:bg-blue-700">
            </form>
        </div>
        <div id="maritime-form" class="tab-content hidden">
            <form method="post" enctype="multipart/form-data" action="/process/maritime">
                <h2 class="text-xl font-semibold mb-4">Upload Maritime PDF and Optional Excel Template</h2>
                <div class="mb-4">
                    <label class="block text-gray-700">PDF File:</label>
                    <input type="file" name="pdf_file" accept=".pdf" class="mt-1 p-2 border rounded w-full" required>
                </div>
                <div class="mb-4">
                    <label class="block text-gray-700">Excel File (optional):</label>
                    <input type="file" name="excel_file" accept=".xlsx" class="mt-1 p-2 border rounded w-full">
                </div>
                <div class="mb-4">
                    <label class="block text-gray-700">Freight Number:</label>
                    <input type="number" name="freight_number" min="1" placeholder="Enter Freight Number (e.g., 200, 250, 500)" class="mt-1 p-2 border rounded w-full">
                </div>
                <div class="mb-4">
                    <label class="block text-gray-700">Container Type:</label>
                    <div class="mt-1">
                        <input type="checkbox" name="container_type" value="40FT" class="mr-2"> 1*40FT<br>
                        <input type="checkbox" name="container_type" value="20FT" class="mr-2"> 1*20FT
                    </div>
                </div>
                <div class="mb-4">
                    <label class="block text-gray-700">Number of Containers:</label>
                    <input type="number" name="num_containers" min="1" placeholder="Enter Number of Containers (default: 1)" class="mt-1 p-2 border rounded w-full">
                </div>
                <input type="submit" value="Upload and Process" class="bg-blue-600 text-white py-2 px-4 rounded hover:bg-blue-700">
            </form>
        </div>
        {% if data %}
        <h2 class="text-xl font-semibold mt-6">Extracted Data</h2>
        <pre class="bg-white p-4 rounded shadow">{{ data | tojson(indent=4) }}</pre>
        {% endif %}
        {% if modified %}
        <h2 class="text-xl font-semibold mt-6">Downloads</h2>
        <a href="/download_excel" class="text-blue-600 hover:underline">Download Modified Excel</a><br>
        <a href="/download_pdf" class="text-blue-600 hover:underline">Download as PDF</a>
        {% endif %}
    </div>
    <script>
        function showTab(tab) {
            document.querySelectorAll('.tab-content').forEach(content => content.classList.add('hidden'));
            document.querySelectorAll('.tab-button').forEach(button => {
                button.classList.remove('text-blue-600', 'border-blue-600');
                button.classList.add('text-gray-600');
            });
            document.getElementById(tab + '-form').classList.remove('hidden');
            document.getElementById(tab + '-tab').classList.add('text-blue-600', 'border-blue-600');
            document.getElementById(tab + '-tab').classList.remove('text-gray-600');
        }
        // Initialize with Normal FERI tab
        showTab('normal');
    </script>
</body>
</html>
'''

def extract_data_from_pdf(pdf_content, pdf_type):
    """
    Extract structured data from the PDF content based on the PDF type (normal or maritime).
    """
    extracted = {}
    with pdfplumber.open(BytesIO(pdf_content)) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text() or ""

    # Clean up text
    text = text.replace('\n', ' ').strip()

    if pdf_type == 'normal':
        # Attestation Number
        match = re.search(r'A\.D\s+N°\s*(\w+)', text)
        if match:
            extracted['attestation_number'] = match.group(1)

        # Importateur (only name)
        match = re.search(r'IMPORTATEUR\s*:\s*([^\s;][^;]*?)(?:\s*;|EXPORTATEUR)', text, re.DOTALL)
        if match:
            importateur_name = match.group(1).strip()
            name_match = re.match(r'^[A-Z\s().]+(?:\s*\([A-Z]+\)\s*[A-Z.]+)?', importateur_name)
            importateur_name = name_match.group(0).strip() if name_match else importateur_name
            extracted['importateur'] = importateur_name
            if importateur_name.endswith(' I'):
                importateur_name = importateur_name[:-2].strip()
            extracted['importateur'] = importateur_name

        # Exporter (only name)
        match = re.search(r'EXPORTATEUR\s*([^\s;][^;]*?)(?:\s*;|TRANSITAIRE)', text, re.DOTALL)
        if match:
            exporter_name = match.group(1).strip()
            name_match = re.match(r'^[A-Z\s()]+(?:\s*\([A-Z]+\)\s*[A-Z]+)?', exporter_name)
            exporter_name = name_match.group(0).strip() if name_match else exporter_name
            if exporter_name.endswith(' E'):
                exporter_name = exporter_name[:-2].strip()
            extracted['exporter'] = exporter_name

        # Forwarding Agent (only name)
        match = re.search(r'TRANSITAIRE\s*:\s*([^\s;][^;]*?)(?:\s*Forwarding agent|DEST\.)', text, re.DOTALL)
        if match:
            forwarding_name = match.group(1).strip()
            name_match = re.match(r'^[A-Z\s().]+(?:\s*\([A-Z]+\)\s*[A-Z.]+)?', forwarding_name)
            extracted['forwarding_agent'] = name_match.group(0).strip() if name_match else forwarding_name

        # Transport ID
        match = re.search(r'TITRE DE TRANSPORT\s*:\s*(.+?)\s*TRANS', text)
        if match:
            extracted['transport_id'] = match.group(1).strip()

        # CBM (Cubic Meters)
        match = re.search(r'(\d+\.\d+\s*CBM)', text)
        if match:
            extracted['cbm'] = match.group(1).strip()

        # Gross Weight
        match = re.search(r'POIDS BRUT\s*:\s*([\d\.]+)\s*Kg', text)
        if match:
            extracted['gross_weight'] = float(match.group(1))

    else:  # maritime
        # FERI Number or Attestation Number
        match = re.search(r'(?:FERI N°|VALIDATION|A\.D\s+N°)\s*:\s*(\w+)', text)
        if match:
            extracted['feri_number'] = match.group(1)

        # Importateur (only name)
        match = re.search(r'IMPORTATEUR\s*:\s*([^\s;][^;]*?)(?:\s*;|EXPORTATEUR|ADD:|$)', text, re.DOTALL)
        if match:
            importateur_name = match.group(1).strip()
            name_match = re.match(r'^[A-Z\s()]+(?:\s*\([A-Z]+\)\s*[A-Z]+)?', importateur_name)
            importateur_name = name_match.group(0).strip() if name_match else importateur_name
            extracted['importateur'] = importateur_name
            if importateur_name.endswith(' I'):
                importateur_name = importateur_name[:-2].strip()
            extracted['importateur'] = importateur_name

        # Transitaire (only name)
        match = re.search(r'TRANSITAIRE\s*:\s*([^\s;][^;]*?)(?:\s*Forwarding agent|DEST\.|ADD:|$)', text, re.DOTALL)
        if match:
            forwarding_name = match.group(1).strip()
            name_match = re.match(r'^[A-Z\s().]+(?:\s*\([A-Z]+\)\s*[A-Z.]+)?', forwarding_name)
            extracted['transitaire'] = name_match.group(0).strip() if name_match else forwarding_name

        # BL (Bill of Lading)
        match = re.search(r'(?:BL|TITRE DE TRANSPORT)\s*:\s*(.+?)(?:\s*ARMATEUR|TRANS|$)', text)
        if match:
            extracted['bl'] = match.group(1).strip()

        # CBM (Cubic Meters)
        match = re.search(r'(\d+\.\d+\s*CBM)\s*(?=(?:POIDS|TEU|CONTENEUR|$))', text)
        if match:
            extracted['cbm'] = match.group(1).strip()

        # Gross Weight
        match = re.search(r'POIDS BRUT\s*:\s*([\d\.]+)\s*(?:Kg|T)', text)
        if match:
            extracted['gross_weight'] = float(match.group(1))

        # Exporter
        match = re.search(r'EXPORTATEUR\s*([^\s;][^;]*?)(?:\s*;|TRANSITAIRE|$)', text, re.DOTALL)
        if match:
            exporter_name = match.group(1).strip()
            name_match = re.match(r'^[A-Z\s().&]+(?:\s*\([A-Z]+\)\s*[A-Z.&]+)?', exporter_name)
            exporter_name = name_match.group(0).strip() if name_match else exporter_name
            if exporter_name.endswith(' E'):
                exporter_name = exporter_name[:-2].strip()
            extracted['exporter'] = exporter_name

    return extracted

@app.route('/')
def index():
    return render_template_string(UPLOAD_FORM, data=None, modified=False)

@app.route('/process/<pdf_type>', methods=['POST'])
def process_pdf(pdf_type):
    if pdf_type not in ['normal', 'maritime']:
        return "Invalid PDF type", 400

    data = None
    modified = False
    if 'pdf_file' not in request.files:
        return "No PDF file part"
    pdf_file = request.files['pdf_file']
    if pdf_file.filename == '':
        return "No selected PDF file"
    if pdf_file and pdf_file.filename.endswith('.pdf'):
        pdf_content = pdf_file.read()
        data = extract_data_from_pdf(pdf_content, pdf_type)

        # Check for Excel file
        excel_content = None
        if 'excel_file' in request.files:
            excel_file = request.files['excel_file']
            if excel_file and excel_file.filename.endswith('.xlsx'):
                excel_content = BytesIO(excel_file.read())

        # Get Freight Number from form input
        freight_number = request.form.get('freight_number', '').strip()
        try:
            freight_number = int(freight_number) if freight_number else None
        except ValueError:
            freight_number = None

        # Get Container Type and Number of Containers (for maritime only)
        container_type = request.form.get('container_type', '') if pdf_type == 'maritime' else ''
        num_containers = 1
        if pdf_type == 'maritime':
            num_containers = request.form.get('num_containers', '').strip()
            try:
                num_containers = int(num_containers) if num_containers else 1
            except ValueError:
                num_containers = 1

        if excel_content:
            # Load workbook
            wb = openpyxl.load_workbook(excel_content)
            ws = wb.active

            # Insert data into specific cells
            if pdf_type == 'normal':
                if 'attestation_number' in data:
                    ws['E6'] = f"FERI/AD: {data['attestation_number']}"
                    ws['B11'] = f"CERTIFICATE (FERI/ADR/AD) No : {data['attestation_number']}"
                if 'forwarding_agent' in data:
                    ws['B8'] = f"DEBTOR: {data['forwarding_agent']}"
                if 'importateur' in data:
                    ws['B10'] = f"IMPORTER: {data['importateur']}"
                if 'transport_id' in data:
                    ws['B14'] = data['transport_id']
                if 'cbm' in data:
                    cbm_value = float(data['cbm'].replace(' CBM', ''))
                    ws['D14'] = cbm_value
                if freight_number is not None:
                    ws['D18'] = freight_number

                fields = ['attestation_number', 'exporter', 'forwarding_agent', 'transport_id', 'cbm', 'gross_weight']
            else:  # maritime
                if 'feri_number' in data:
                    ws['E6'] = f"FERI/AD: {data['feri_number']}"
                    ws['B11'] = f"CERTIFICATE (FERI/ADR/AD) No : {data['feri_number']}"
                if 'transitaire' in data:
                    ws['B8'] = f"DEBTOR: {data['transitaire']}"
                if 'importateur' in data:
                    ws['B10'] = f"IMPORTER: {data['importateur']}"
                if 'bl' in data:
                    ws['B14'] = data['bl']
                if 'cbm' in data and container_type not in ['40FT', '20FT']:
                    cbm_value = float(data['cbm'].replace(' CBM', ''))
                    ws['D14'] = cbm_value
                if container_type == '40FT':
                    ws['D14'] = 110
                elif container_type == '20FT':
                    ws['D14'] = 60
                ws['E14'] = num_containers
                ws['D17'] = num_containers
                ws['D18'] = num_containers * 250
                if freight_number is not None:
                    ws['D18'] = freight_number

                fields = ['feri_number', 'exporter', 'transitaire', 'bl', 'cbm', 'gross_weight']

            # Append row data
            row_data = [data.get(f, '') for f in fields]
            ws.append(row_data)

            # Save modified Excel
            modified_excel = BytesIO()
            wb.save(modified_excel)
            modified_excel.seek(0)

            # Create PDF from the Excel data
            sheet_data = [list(row) for row in ws.iter_rows(values_only=True)]
            pdf_buffer = BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
            table = Table(sheet_data)
            doc.build([table])
            pdf_buffer.seek(0)

            # Store in globals
            global modified_excel_global, modified_pdf_global
            modified_excel_global = modified_excel
            modified_pdf_global = pdf_buffer

            modified = True

    return render_template_string(UPLOAD_FORM, data=data, modified=modified)

@app.route('/download_excel')
def download_excel():
    global modified_excel_global
    if modified_excel_global is None:
        return "No modified Excel available"
    modified_excel_global.seek(0)
    return send_file(modified_excel_global, as_attachment=True, download_name='modified.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/download_pdf')
def download_pdf():
    global modified_pdf_global
    if modified_pdf_global is None:
        return "No PDF available"
    modified_pdf_global.seek(0)
    return send_file(modified_pdf_global, as_attachment=True, download_name='modified.pdf', mimetype='application/pdf')

if __name__ == '__main__':
    app.run(debug=True)