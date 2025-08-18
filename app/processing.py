import openpyxl
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table
import os
import glob
import json

# Directory containing Excel templates
laban_dir = 'template/laban'  # Directory for Normal FERI templates
malaba_dir = 'template/malaba'  # Directory for Maritime templates

# Global variables to store the modified files and data in memory
modified_excel_global = None
modified_pdf_global = None
data_global = None
pdf_type_global = None

def get_available_templates(template_dir):
    """Get list of available Excel templates from the specified directory"""
    if not os.path.exists(template_dir):
        os.makedirs(template_dir)
        return []
    
    excel_files = glob.glob(os.path.join(template_dir, '*.xlsx'))
    return [os.path.basename(f) for f in excel_files]

def process_excel_and_pdf(data, pdf_type, template_file, freight_number, container_type='', num_containers=1):
    """
    Process extracted data, update Excel template, and generate PDF.
    Returns modified Excel, PDF buffers, and formatted JSON data.
    """
    global modified_excel_global, modified_pdf_global, data_global, pdf_type_global
    json_data = json.dumps(data, ensure_ascii=False, indent=2)
    modified_excel = None
    modified_pdf = None

    # Store data and pdf_type globally
    data_global = data
    pdf_type_global = pdf_type

    # Check for selected template
    excel_content = None
    if template_file:
        template_path = os.path.join(laban_dir if pdf_type == 'normal' else malaba_dir, template_file)
        if os.path.exists(template_path):
            with open(template_path, 'rb') as f:
                excel_content = BytesIO(f.read())
        template_path = os.path.join(laban_dir if pdf_type == 'normal' else malaba_dir, template_file)
        if os.path.exists(template_path):
            with open(template_path, 'rb') as f:
                excel_content = BytesIO(f.read())

    if excel_content:
        # Load workbook
        wb = openpyxl.load_workbook(excel_content)
        ws = wb.active

        # Insert data into specific cells
        if pdf_type == 'normal':
            if 'attestation_number' in data:
                ws['E6'] = f"FERI/AD: {data['attestation_number']}"
                ws['B11'] = f"CERTIFICATE (FERI/ADR/AD) No : {data['attestation_number']}"
            if 'forwarding_agent' in data:
                ws['B8'] = f"DEBTOR: {data['forwarding_agent']}"
            if 'importateur' in data:
                ws['B10'] = f"IMPORTER: {data['importateur']}"
            if 'transport_id' in data:
                ws['B14'] = data['transport_id']
            if 'cbm' in data:
                cbm_value = float(data['cbm'].replace(' CBM', ''))
                ws['D14'] = cbm_value
            if freight_number is not None:
                ws['D18'] = freight_number

            fields = ['attestation_number', 'exporter', 'forwarding_agent', 'transport_id', 'cbm', 'gross_weight']
        else:  # maritime
            if 'feri_number' in data:
                ws['E6'] = f"FERI/AD: {data['feri_number']}"
                ws['B11'] = f"CERTIFICATE (FERI/ADR/AD) No : {data['feri_number']}"
            if 'transitaire' in data:
                ws['B8'] = f"DEBTOR: {data['transitaire']}"
            if 'importateur' in data:
                ws['B10'] = f"IMPORTER: {data['importateur']}"
            if 'bl' in data:
                ws['B14'] = data['bl']
            if 'cbm' in data and container_type not in ['40FT', '20FT']:
                cbm_value = float(data['cbm'].replace(' CBM', ''))
                ws['D14'] = cbm_value
            if container_type == '40FT':
                ws['D14'] = 110
            elif container_type == '20FT':
                ws['D14'] = 60
            ws['E14'] = num_containers
            ws['D17'] = num_containers
            ws['D18'] = num_containers * 250
            if freight_number is not None:
                ws['D18'] = freight_number

            fields = ['feri_number', 'exporter', 'transitaire', 'bl', 'cbm', 'gross_weight']

        # Append row data
        row_data = [data.get(f, '') for f in fields]
        ws.append(row_data)

        # Save modified Excel
        modified_excel = BytesIO()
        wb.save(modified_excel)
        modified_excel.seek(0)

        # Create PDF from the Excel data
        sheet_data = [list(row) for row in ws.iter_rows(values_only=True)]
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        table = Table(sheet_data)
        doc.build([table])
        pdf_buffer.seek(0)

        # Store in globals
        global modified_excel_global, modified_pdf_global
        modified_excel_global = modified_excel
        modified_pdf_global = pdf_buffer

    return modified_excel, modified_pdf, json_data